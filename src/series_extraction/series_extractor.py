from objects import (
    Worksheet,
    WorkbookData,
    Table,
    Cell,
    Series,
    SeriesId,
    SeriesDataType,
)
from openpyxl.utils import get_column_letter

from typing import Dict, List, Tuple


class SeriesExtractor:

    @staticmethod
    def build_series(
        workbook_data: WorkbookData,
        sheet: Worksheet,
        header: str,
        start_row_or_column: int,
        end_row_or_column: int,
        index: int,
    ) -> Series:
        """Build a series object."""
        start_cell_row = start_row_or_column + 1
        start_cell_column = index

        row_formulas, row_values = [], []
        sheet_data = workbook_data.get_sheet_data(sheet.sheet_name)
        for offset in range(1, 3):
            cell_key = f"{get_column_letter(index)}{start_row_or_column + offset}"
            cell = sheet_data.get(cell_key)
            row_formulas.append(cell.formula if cell else None)
            row_values.append(cell.value if cell else None)

        series_header_cell_row, series_header_cell_column = (
            SeriesExtractor.calculate_header_cell(start_cell_row, start_cell_column)
        )

        series_id = SeriesId(
            sheet_name=sheet.sheet_name,
            series_header=header,
            series_header_cell_row=series_header_cell_row,
            series_header_cell_column=series_header_cell_column,
        )

        series_data_type = (
            SeriesDataType(type(row_values[0]).__name__)
            if row_values
            else SeriesDataType.STR
        )

        return Series(
            series_id=series_id,
            worksheet=sheet,
            series_header=header,
            formulas=row_formulas,
            values=row_values,
            series_starting_cell=Cell(
                row=start_cell_row,
                column=start_cell_column,
            ),
            series_length=end_row_or_column - start_row_or_column,
            series_data_type=series_data_type,
        )

    @staticmethod
    def handle_series(
        workbook_data: WorkbookData,
        sheet: Worksheet,
        header_values: List[str],
        start_column: int,
        end_column: int,
        start_row: int,
        end_row: int,
    ) -> Dict[str, Series]:
        """Handle the series extraction."""
        series_data = {}
        for index, header in enumerate(header_values, start=start_column):
            range_identifier = f"{get_column_letter(index)}{start_row}:{get_column_letter(index)}{end_row}"
            series = SeriesExtractor.build_series(
                workbook_data,
                sheet,
                header,
                start_row,
                end_row,
                index,
            )
            series_data[range_identifier] = series
        return series_data

    @staticmethod
    def extract_table_details(
        extracted_tables: Dict[Worksheet, List[Table]], workbook_data: WorkbookData
    ) -> Dict[Worksheet, Dict[str, Series]]:
        """Extract series from the tables."""
        tables_data = {}
        for sheet, tables in extracted_tables.items():
            sheet_data = {}
            for table in tables:
                series_result = SeriesExtractor.handle_series(
                    workbook_data,
                    sheet,
                    table.header_values,
                    table.range.start_cell.column,
                    table.range.end_cell.column,
                    table.range.start_cell.row,
                    table.range.end_cell.row,
                )
                sheet_data.update(series_result)
            tables_data[sheet] = sheet_data
        return tables_data

    @staticmethod
    def calculate_header_cell(
        series_starting_cell_row: int,
        series_starting_cell_column: int,
    ) -> Tuple[int, int]:
        """Calculate the header cell for the series."""
        return (series_starting_cell_row - 1, series_starting_cell_column)

    @staticmethod
    def extract_series(
        extracted_tables: Dict[Worksheet, List[Table]],
        workbook_data: WorkbookData,
    ) -> Dict[str, List[Series]]:
        """Extract series from the tables."""
        detailed_series = SeriesExtractor.extract_table_details(
            extracted_tables, workbook_data
        )

        series_collection = {}

        for worksheet, table_data in detailed_series.items():
            series_collection[worksheet.sheet_name] = []
            for _, single_series in table_data.items():
                series_collection[worksheet.sheet_name].append(single_series)

        return series_collection
