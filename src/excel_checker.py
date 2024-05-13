import openpyxl
import math


class ExcelChecker:
    @staticmethod
    def excels_are_equivalent(file1_path: str, file2_path: str):
        # Load the workbooks
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)

        # Check if the number of sheets is the same
        if len(wb1.sheetnames) != len(wb2.sheetnames):
            print("Number of sheets is different")
            return False

        # Create dictionaries to store sheet data
        wb1_sheets = {sheet: wb1[sheet].values for sheet in wb1.sheetnames}
        wb2_sheets = {sheet: wb2[sheet].values for sheet in wb2.sheetnames}

        # Check if the sheet names are the same
        if set(wb1_sheets.keys()) != set(wb2_sheets.keys()):
            print("Sheet names are different")
            return False

        # Compare the values in each sheet
        for sheet in wb1_sheets:
            wb1_values = list(wb1_sheets[sheet])
            wb2_values = list(wb2_sheets[sheet])

            for row1, row2 in zip(wb1_values, wb2_values):
                if row1 != row2:
                    for cell1, cell2 in zip(row1, row2):
                        if type(cell1) == float and type(cell2) == float:
                            if not math.isclose(cell1, cell2, abs_tol=0.0001):
                                print(
                                    f"Numerical values in sheet '{sheet}' differ more than 0.0001 tolerance."
                                )
                                return False
                        elif cell1 != cell2:
                            print(f"Values in sheet '{sheet}' are different")
                            print(f"wb1: {cell1}")
                            print(f"wb2: {cell2}")
                            return False

        # If all checks pass, the workbooks are equivalent
        return True
