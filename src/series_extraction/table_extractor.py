from objects import (
    ExcelFile,
    Worksheet,
    Table,
    LocatedTables,
    HeaderLocation,
    Cell,
    CellRange,
    WorkbookData
)
from excel_utils import ExcelUtils
from typing import Dict, List, Set, Tuple, Optional, Union
import openpyxl


class CellOperations:

    @staticmethod
    def _cells_are_adjacent(cell1: Cell, cell2: Cell) -> bool:
        """Check if two cells are adjacent based on their row and column indices."""
        return abs(cell1.row - cell2.row) <= 1 and abs(cell1.column - cell2.column) <= 1

    @staticmethod
    def _extract_non_empty_cells(sheet_data: Dict) -> Set[Cell]:
        """Extract non-empty cells from the sheet data."""
        return {cell for cell in sheet_data.values() if cell.value is not None}


class TableLocator:

    @staticmethod
    def _process_adjacent_cells(
        frontier: Set[Cell], non_empty_cells: Set[Cell]
    ) -> Set[Cell]:
        """Identify and process all adjacent cells for the given frontier."""
        new_frontier = set()
        for frontier_cell in frontier:
            adjacent_cells = {
                cell
                for cell in non_empty_cells
                if CellOperations._cells_are_adjacent(cell, frontier_cell)
            }
            non_empty_cells.difference_update(adjacent_cells)
            new_frontier.update(adjacent_cells)
        return new_frontier

    @staticmethod
    def _update_cell_cluster(
        cluster: Set[Cell], new_frontier: Set[Cell]
    ) -> Tuple[Set[Cell], int, int, int, int]:
        """Update the cluster with new cells and adjust the boundaries."""
        cluster.update(new_frontier)
        min_row = min(cluster, key=lambda x: x.row).row
        max_row = max(cluster, key=lambda x: x.row).row
        min_col = min(cluster, key=lambda x: x.column).column
        max_col = max(cluster, key=lambda x: x.column).column
        return cluster, min_row, max_row, min_col, max_col

    @staticmethod
    def _expand_cell_cluster(
        frontier: Set[Cell], non_empty_cells: Set[Cell]
    ) -> Tuple[Set[Cell], int, int, int, int]:
        """Expand the cluster from the frontier cells, updating boundaries."""
        cluster = set(frontier)
        while frontier:
            new_frontier = TableLocator._process_adjacent_cells(
                frontier, non_empty_cells
            )
            cluster, min_row, max_row, min_col, max_col = (
                TableLocator._update_cell_cluster(cluster, new_frontier)
            )
            frontier = new_frontier

        return cluster, min_row, max_row, min_col, max_col

    @staticmethod
    def _find_table_boundaries(
        sheet_name: str, sheet_data: Dict[str, Cell]
    ) -> List[Table]:
        """Identify table boundaries by clustering adjacent non-empty cells and return as Table objects."""
        non_empty_cells = CellOperations._extract_non_empty_cells(sheet_data)
        tables = []

        while non_empty_cells:
            initial_cell = non_empty_cells.pop()
            cluster, min_row, max_row, min_col, max_col = (
                TableLocator._expand_cell_cluster({initial_cell}, non_empty_cells)
            )
            table = Table(
                name=f"{sheet_name}_{len(tables) + 1}",
                range=CellRange(
                    start_cell=Cell(
                        column=min_col,
                        row=min_row,
                        sheet_name=sheet_name,
                    ),
                    end_cell=Cell(
                        column=max_col,
                        row=max_row,
                        sheet_name=sheet_name,
                    ),
                ),
            )
            tables.append(table)

        return tables

    @staticmethod
    def _locate_data_tables(workbook_data: WorkbookData) -> List[LocatedTables]:
        """Locate tables in the given data stored in WorkbookData."""
        all_located_tables = []
        for sheet_name, sheet_data in workbook_data.data.items():
            tables = TableLocator._find_table_boundaries(sheet_name, sheet_data)
            located_tables = LocatedTables(
                worksheet=Worksheet(sheet_name=sheet_name), tables=tables
            )
            all_located_tables.append(located_tables)
        return all_located_tables


class DataExtractor:

    @staticmethod
    def _are_first_row_values_strings(
        range_input: CellRange, data_object: Dict[str, Cell]
    ) -> Tuple[bool, Optional[List[Union[int, str, float, bool]]]]:

        start_cell = range_input.start_cell
        end_cell = range_input.end_cell

        header_values = []

        for column in range(start_cell.column, end_cell.column + 1):
            cell_coordinate = (
                f"{ExcelUtils.get_column_letter_from_number(column)}{start_cell.row}"
            )
            if data_object[cell_coordinate].value_type != "str":
                return False, None
            else:
                header_values.append(data_object[cell_coordinate].value)

        return True, header_values

    @staticmethod
    def _get_first_column_values(
        range_input: CellRange, data_object: Dict[str, Cell]
    ) -> List[Union[int, str, float, bool]]:
        """Get values from the first column of a range."""
        start_cell = range_input.start_cell
        end_cell = range_input.end_cell

        first_column_values = []

        for row in range(start_cell.row, end_cell.row + 1):
            cell_coordinate = (
                f"{ExcelUtils.get_column_letter_from_number(start_cell.column)}{row}"
            )
            first_column_values.append(data_object[cell_coordinate].value)

        return first_column_values

    @staticmethod
    def _get_header_location_and_values(
        workbook_data: WorkbookData,
    ) -> List[LocatedTables]:
        """Get header location and values for the located tables."""
        located_tables = TableLocator._locate_data_tables(workbook_data)
        for located_table in located_tables:
            sheet_data = workbook_data.get_sheet_data(
                located_table.worksheet.sheet_name
            )
            for table in located_table.tables:
                boolean, header_values = DataExtractor._are_first_row_values_strings(
                    table.range, sheet_data
                )
                if boolean:
                    table.header_location = HeaderLocation.TOP
                    table.header_values = header_values
                else:
                    table.header_location = HeaderLocation.LEFT
                    table.header_values = DataExtractor._get_first_column_values(
                        table.range, sheet_data
                    )
        return located_tables


class TableExtractor:

    @staticmethod
    def extract_tables(
        excel_file: ExcelFile,
    ) -> Tuple[Dict[Worksheet, List[Table]], WorkbookData]:
        """Extract tables from the given Excel file and return as a dictionary of tables and the workbook cell data."""
        workbook_data = WorkbookData()
        for ws_values, ws_formulas in zip(
            excel_file.workbook_with_values.worksheets,
            excel_file.workbook_with_formulas.worksheets,
        ):
            worksheet = Worksheet(sheet_name=ws_values.title, worksheet=ws_values)
            sheet_data = TableExtractor._extract_sheet_data(ws_values, ws_formulas)
            workbook_data.add_sheet_data(worksheet.sheet_name, sheet_data)

        located_tables = DataExtractor._get_header_location_and_values(workbook_data)

        extracted_tables = {}
        for located_table in located_tables:
            extracted_tables[
                Worksheet(sheet_name=located_table.worksheet.sheet_name)
            ] = located_table.tables

        return extracted_tables, workbook_data

    @staticmethod
    def _extract_cell_data(
        cell_with_value: openpyxl.cell.cell.Cell,
        cell_with_formula: openpyxl.cell.cell.Cell,
    ) -> Cell:
        """Extract data from a cell and return as a Cell object."""
        formula = (
            cell_with_formula.value
            if isinstance(cell_with_formula.value, str)
            and cell_with_formula.value.startswith("=")
            else None
        )
        return Cell(
            column=cell_with_value.column,
            row=cell_with_value.row,
            coordinate=cell_with_value.coordinate,
            sheet_name=cell_with_value.parent.title,
            value=cell_with_value.value,
            value_type=type(cell_with_value.value).__name__,
            formula=formula,
        )

    @staticmethod
    def _extract_sheet_data(
        worksheet_with_values: Worksheet, worksheet_with_formulas: Worksheet
    ) -> Dict[str, Cell]:
        """Extract data from a worksheet and return as a dictionary of cells."""
        sheet_data = {}
        for row_values, row_formulas in zip(
            worksheet_with_values.iter_rows(), worksheet_with_formulas.iter_rows()
        ):
            for cell_with_value, cell_with_formula in zip(row_values, row_formulas):
                cell = TableExtractor._extract_cell_data(
                    cell_with_value, cell_with_formula
                )
                sheet_data[cell.coordinate] = cell
        return sheet_data
