import openpyxl
from typing import Dict, List
from objects import Series


class ExcelBuilder:
    @staticmethod
    def create_excel_from_workbook(
        workbook: openpyxl.Workbook, output_file_path: str
    ) -> None:
        """Creates a new Excel workbook by duplicating all sheets and their cell contents from the given workbook."""
        new_workbook = openpyxl.Workbook()
        ExcelBuilder._remove_default_sheet(new_workbook)
        ExcelBuilder._copy_sheets_and_contents(workbook, new_workbook)
        new_workbook.save(output_file_path)

    @staticmethod
    def create_excel_from_series(
        series_collection: List[Series],
        output_file_path: str,
        values_only: bool = False,
    ) -> None:
        """Create an Excel file from a list of series, optionally using only values or formulas."""
        new_workbook = openpyxl.Workbook()
        ExcelBuilder._remove_default_sheet(new_workbook)
        worksheets_dictionary: Dict[str, openpyxl.worksheet.worksheet.Worksheet] = {}

        for series_item in series_collection:
            current_worksheet = ExcelBuilder._get_or_create_worksheet(
                new_workbook, series_item, worksheets_dictionary
            )
            ExcelBuilder._place_series_header(current_worksheet, series_item)
            ExcelBuilder._fill_cells_with_data(
                current_worksheet, series_item, values_only
            )

        new_workbook.save(output_file_path)

    @staticmethod
    def _remove_default_sheet(workbook: openpyxl.Workbook) -> None:
        """Removes the default sheet from a new workbook to start with a clean slate."""
        workbook.remove(workbook.active)

    @staticmethod
    def _get_or_create_worksheet(
        workbook: openpyxl.Workbook,
        series: Series,
        worksheets_dictionary: Dict[str, openpyxl.worksheet.worksheet.Worksheet],
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        """Get an existing worksheet by name or create a new one if it doesn't exist."""
        sheet_name = series.series_id.sheet_name
        if sheet_name not in worksheets_dictionary:
            ws = (
                workbook.create_sheet(title=sheet_name)
                if sheet_name not in workbook.sheetnames
                else workbook[sheet_name]
            )
            worksheets_dictionary[sheet_name] = ws
        return worksheets_dictionary[sheet_name]

    @staticmethod
    def _place_series_header(
        worksheet: openpyxl.worksheet.worksheet.Worksheet, series: Series
    ) -> None:
        """Place the header for a series in the specified worksheet."""
        header_row = series.series_id.series_header_cell_row
        header_col = series.series_id.series_header_cell_column
        series_header = series.series_header
        worksheet.cell(row=header_row, column=header_col, value=series_header)

    @staticmethod
    def _fill_cells_with_data(
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        series: Series,
        values_only: bool,
    ) -> None:
        """Fill worksheet cells with data from a series, either formulas or values."""
        start_row = series.series_starting_cell.row
        start_col = series.series_starting_cell.column
        header_location = series.header_location.value
        formulas = series.formulas
        values = series.values

        for i in range(series.series_length):
            row = start_row + i if header_location == "top" else start_row
            col = start_col + i if header_location == "left" else start_col
            cell = worksheet.cell(row=row, column=col)
            if not values_only:
                cell.value = formulas[i] if formulas != [None, None] else values[i]
            else:
                cell.value = values[i]

    @staticmethod
    def _copy_sheets_and_contents(
        source_workbook: openpyxl.Workbook, target_workbook: openpyxl.Workbook
    ) -> None:
        """Copies all sheets and their cell contents from the source workbook to the target workbook."""
        for sheet in source_workbook:
            new_sheet = target_workbook.create_sheet(title=sheet.title)
            for row in sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet.cell(row=cell.row, column=cell.column)
                    new_cell.value = cell.value
